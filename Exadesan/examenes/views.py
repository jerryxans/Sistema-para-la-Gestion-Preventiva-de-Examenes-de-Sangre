from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.http import HttpResponse
from django.contrib.auth import logout
from django.contrib.auth.forms import AuthenticationForm
import json
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponseForbidden
from rest_framework import viewsets, status, filters as drf_filters
from rest_framework.decorators import action
from rest_framework.response import Response
from django.http import FileResponse, Http404
from .models import ResultadoExamen, Perfil, CatalogoExamen
from .serializers import ExamenSerializer
from .permissions import EsPropietarioOMedicoAdmin
from django.http import FileResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from io import BytesIO
from .models import TipoExamen, Perfil
from .models import CatalogoExamen
from .forms import CatalogoExamenForm
from django.contrib import messages
from django.contrib.auth.forms import UserChangeForm
from django.contrib.auth import authenticate, login
from .forms import ResultadoExamenForm
from datetime import datetime
from reportlab.lib import colors
from textwrap import wrap
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import cm
from .forms import UsuarioForm
from django.dispatch import receiver
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt


# HOME
def home(request):
    return render(request, "examenes/home.html")

@login_required
def redirect_dashboard(request):
    try:
        perfil = request.user.perfil
    except Perfil.DoesNotExist:
        return redirect('examenes:home')

    if perfil.rol == 'ADMIN':
        return redirect('examenes:dashboard_admin')
    elif perfil.rol == 'MEDICO':
        return redirect('examenes:dashboard_medico')
    elif perfil.rol == 'PACIENTE':
        return redirect('examenes:dashboard_paciente')
    else:
        return redirect('examenes:home')

# Vista de login
def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            return redirect('examenes:redirect_dashboard')
        else:
            messages.error(request, "Usuario o contraseña incorrectos.")
            return redirect('login')

    return render(request, 'examenes/login.html')


# Vista central de redirección
@login_required
def redirect_dashboard(request):
    try:
        perfil = request.user.perfil
    except Perfil.DoesNotExist:
        return redirect('examenes:home')

    if perfil.rol == 'ADMIN':
        return redirect('examenes:dashboard_admin')
    elif perfil.rol == 'MEDICO':
        return redirect('examenes:dashboard_medico')
    elif perfil.rol == 'PACIENTE':
        return redirect('examenes:dashboard_paciente')
    else:
        return redirect('examenes:home')


# Ejemplo de vistas de cada dashboard
@login_required
def dashboard_admin(request):
    messages.info(request, "👨‍⚕️ Bienvenido, Administrador.")
    return render(request, "examenes/dashboard_admin.html")

def dashboard_medico(request):
    messages.info(request, "🩺 Bienvenido Doctor.")
    return render(request, "examenes/dashboard_medico.html")

def dashboard_paciente(request):
    messages.info(request, "📊 Estimado paciente, bienvenido, aca podra visualizar sus examenes, graficos, resultados y mas.")
    return render(request, "examenes/dashboard_paciente.html")

def custom_logout(request):
    logout(request)  # Cierra la sesión
    return redirect('examenes:redirect_dashboard')  # Redirige al home

def login_view(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(request, username=username, password=password)

            if user is not None:
                login(request, user)
                try:
                    perfil = user.perfil
                    if perfil.rol == 'PACIENTE':
                        messages.success(request, "✅ Bienvenido, paciente.")
                        return redirect('examenes:dashboard_paciente')
                    elif perfil.rol == 'MEDICO':
                        messages.success(request, "✅ Bienvenido, médico.")
                        return redirect('examenes:dashboard_medico')
                    elif perfil.rol == 'ADMIN':
                        messages.success(request, "✅ Bienvenido, administrador.")
                        return redirect('admin:index')
                except Perfil.DoesNotExist:
                    messages.error(request, "❌ No se encontró perfil asociado a este usuario.")
                    return redirect('login')
            else:
                messages.error(request, "❌ Usuario o contraseña incorrectos.")
        else:
            messages.error(request, "⚠️ Revise los campos ingresados.")
    else:
        form = AuthenticationForm()

    return render(request, 'login.html', {'form': form})

# --- LISTA DE EXÁMENES ---
@login_required
def lista_examenes(request):
    perfil, _ = Perfil.objects.get_or_create(
        usuario=request.user,
        defaults={'rol': 'PACIENTE'}
    )
    paciente_id = request.GET.get("paciente")

    # PACIENTE: solo sus exámenes
    if perfil.rol == 'PACIENTE':
        examenes = ResultadoExamen.objects.filter(paciente=request.user).order_by('-fecha_examen')
        context = {"examenes": examenes, "paciente_actual": request.user}

    # MÉDICO o ADMIN: exámenes del paciente seleccionado
    elif perfil.rol in ('MEDICO', 'ADMIN'):
        if paciente_id:
            examenes = ResultadoExamen.objects.filter(paciente_id=paciente_id).order_by('-fecha_examen')
            paciente_obj = User.objects.filter(id=paciente_id).first()
            context = {"examenes": examenes, "paciente_actual": paciente_obj}
        else:
            examenes = []
            context = {
                "examenes": examenes,
                "mensaje": "Selecciona un paciente desde la lista para ver sus exámenes."
            }

    else:
        examenes = []
        context = {"examenes": examenes}

    return render(request, "examenes/lista_examenes.html", context)


# --- CARGA DE EXÁMENES ---
@login_required
def cargar_examen(request):
    if request.method == 'POST':
        form = ResultadoExamenForm(request.POST, request.FILES)
        if form.is_valid():
            examen = form.save()

            # Si la petición es AJAX responder JSON
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({
                    "success": True,
                    "id": examen.id,
                    "fecha_examen": examen.fecha_examen.strftime("%d/%m/%Y"),
                    "examen_tipo": examen.examen_tipo.nombre,
                    "valor": str(examen.valor),
                    "nivel_riesgo": examen.nivel_riesgo
                })

            messages.success(request, "✅ Examen cargado exitosamente.")
            return redirect('examenes:cargar_examen')
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({"success": False, "errors": form.errors})

            messages.error(request, "❌ Revisa los datos ingresados.")
    else:
        form = ResultadoExamenForm()

    return render(request, 'examenes/cargar_examen.html', {'form': form})





# --- LISTA DE PACIENTES ---
@login_required
def lista_pacientes(request):
    pacientes = User.objects.filter(perfil__rol='PACIENTE')
    return render(request, 'examenes/lista_pacientes.html', {'pacientes': pacientes})

@login_required
def ver_examen(request, examen_id):
    examen = get_object_or_404(ResultadoExamen, id=examen_id)

    # Control de acceso, solo el paciente dueño o un médico/admin pueden verlo
    perfil, _ = Perfil.objects.get_or_create(usuario=request.user, defaults={'rol': 'PACIENTE'})
    if perfil.rol == 'PACIENTE' and examen.paciente != request.user:
        return HttpResponseForbidden("No tienes permiso para ver este examen.")

    return render(request, "examenes/ver_examen.html", {"examen": examen})

@login_required
def descargar_examen(request, examen_id):
    examen = get_object_or_404(ResultadoExamen, id=examen_id)

    # Control de acceso, solo el paciente dueño o un médico/admin pueden descargar
    perfil, _ = Perfil.objects.get_or_create(usuario=request.user, defaults={'rol': 'PACIENTE'})
    if perfil.rol == 'PACIENTE' and examen.paciente != request.user:
        return HttpResponseForbidden("No tienes permiso para descargar este examen.")

    if not examen.documento_pdf:
        raise Http404("Este examen no tiene documento adjunto.")

    return FileResponse(examen.documento_pdf.open(), as_attachment=True, filename=examen.documento_pdf.name)

@login_required
def descargar_examen(request, examen_id):
    examen = get_object_or_404(ResultadoExamen, id=examen_id)
    perfil, _ = Perfil.objects.get_or_create(usuario=request.user, defaults={'rol': 'PACIENTE'})

    if perfil.rol == 'PACIENTE' and examen.paciente != request.user:
        return HttpResponseForbidden("No tienes permiso para descargar este examen.")

    # Si hay archivo adjunto lo devuelve
    if examen.documento_pdf:
        return FileResponse(examen.documento_pdf.open(), as_attachment=True, filename=examen.documento_pdf.name)

    # Si no hay archivo, genera PDF dinámico
    buffer = BytesIO()
    p = canvas.Canvas(buffer)
    p.setFont("Helvetica", 12)
    p.drawString(100, 800, f"Resultado de Examen Clínico")
    p.drawString(100, 780, f"Paciente: {examen.paciente.username}")
    p.drawString(100, 760, f"Tipo de examen: {examen.examen_tipo.nombre}")
    p.drawString(100, 740, f"Fecha: {examen.fecha_examen.strftime('%d/%m/%Y')}")
    p.drawString(100, 720, f"Resultado: {examen.valor}")
    p.drawString(100, 700, f"Nivel de riesgo: {examen.nivel_riesgo}")
    p.showPage()
    p.save()
    buffer.seek(0)

    return FileResponse(buffer, as_attachment=True, filename=f"examen_{examen.id}.pdf")

# ---  DESCARGAR REPORTES EN EXCEL---
def riesgo_color(riesgo: str):
    r = (riesgo or "").strip().lower()
    if r == "normal":
        return "92D050"
    if r == "alto":
        return "FF0000"
    if r == "bajo":
        return "FFC000"
    return "FFFFFF"

@login_required
def descargar_reportes_excel(request):
    paciente = request.user
    examenes = ResultadoExamen.objects.filter(paciente=paciente).order_by('-fecha_examen')

    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados Clínicos"

    # Encabezados
    headers = ["Fecha", "Examen", "Valor", "Riesgo", "Observaciones"]
    ws.append(headers)

    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Filas de resultados
    row_num = 2
    for examen in examenes:
        ws.cell(row=row_num, column=1, value=examen.fecha_examen.strftime("%d/%m/%Y"))
        ws.cell(row=row_num, column=2, value=examen.examen_tipo.nombre)
        ws.cell(row=row_num, column=3, value=examen.valor)

        riesgo_cell = ws.cell(row=row_num, column=4, value=examen.nivel_riesgo)
        riesgo_cell.fill = PatternFill(start_color=riesgo_color(examen.nivel_riesgo),
                                       end_color=riesgo_color(examen.nivel_riesgo),
                                       fill_type="solid")
        riesgo_cell.alignment = Alignment(horizontal="center")
        riesgo_cell.font = Font(bold=True)

        ws.cell(row=row_num, column=5, value=examen.observaciones_medline or "")

        row_num += 1

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=reportes_{paciente.username}.xlsx'
    wb.save(response)
    return response

# ---  DESCARGAR REPORTES EN PDF---

def riesgo_color_html(riesgo: str) -> str:
    r = (riesgo or "").strip().lower()
    if r == "normal":
        return f'<font color="green">{riesgo}</font>'
    if r == "alto":
        return f'<font color="red">{riesgo}</font>'
    if r == "bajo":
        return f'<font color="orange">{riesgo}</font>'
    return f'{riesgo}'

@login_required
def descargar_reportes_pdf(request):
    paciente = request.user
    examenes = ResultadoExamen.objects.filter(paciente=paciente).order_by('-fecha_examen')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=reportes_{paciente.username}.pdf'

    doc = SimpleDocTemplate(
        response,
        pagesize=letter,
        leftMargin=2*cm,
        rightMargin=2*cm,
        topMargin=2.2*cm,
        bottomMargin=2*cm,
    )

    styles = getSampleStyleSheet()
    # Encabezados
    title_style = ParagraphStyle(
        'TitleClinical',
        parent=styles['Title'],
        fontName='Helvetica-Bold',
        fontSize=16,
        leading=20,
        alignment=1,
    )
    meta_style = ParagraphStyle(
        'Meta',
        parent=styles['Normal'],
        fontSize=11,
        leading=14,
        spaceAfter=6,
    )
    cell_style = ParagraphStyle(
        'Cell',
        parent=styles['Normal'],
        fontSize=9,
        leading=12,
    )
    header_cell_style = ParagraphStyle(
        'HeaderCell',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=10,
        leading=12,
    )

    elements = []
    elements.append(Paragraph("Reportes Clínicos", title_style))
    elements.append(Paragraph(f"Paciente: {paciente.username}", meta_style))
    elements.append(Paragraph(f"Fecha: {datetime.now().strftime('%d/%m/%Y')}", meta_style))
    elements.append(Spacer(1, 8))

    # Cabecera de la tabla
    data = [
        [
            Paragraph("Fecha", header_cell_style),
            Paragraph("Examen", header_cell_style),
            Paragraph("Valor", header_cell_style),
            Paragraph("Riesgo", header_cell_style),
            Paragraph("Observaciones", header_cell_style),
        ]
    ]

    # Filas
    for ex in examenes:
        fecha = Paragraph(ex.fecha_examen.strftime('%d/%m/%Y'), cell_style)
        examen = Paragraph(ex.examen_tipo.nombre, cell_style)
        valor = Paragraph(str(ex.valor), cell_style)
        riesgo = Paragraph(riesgo_color_html(ex.nivel_riesgo), cell_style)
        observaciones = Paragraph(ex.observaciones_medline or "", cell_style)

        data.append([fecha, examen, valor, riesgo, observaciones])

    # Anchos de columnas
    col_widths = [3*cm, 5*cm, 2.5*cm, 3*cm, 7*cm]

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        # Cabecera
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#F0F3F7')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor('#000000')),
        ('LINEBELOW', (0,0), (-1,0), 0.75, colors.HexColor('#D0D7DF')),
        ('FONT', (0,0), (-1,0), 'Helvetica-Bold', 10),

        # Bordes y líneas
        ('GRID', (0,0), (-1,-1), 0.25, colors.HexColor('#DDE3EA')),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING', (0,0), (-1,-1), 6),
        ('RIGHTPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),

        # Zebra
        ('BACKGROUND', (0,1), (-1,-1), colors.white),
    ]))

    elements.append(table)
    doc.build(elements)
    return response



@login_required
def lista_usuarios(request):
    perfil = Perfil.objects.get(usuario=request.user)

    # Solo el ADMIN puede ver todos los usuarios
    if perfil.rol != 'ADMIN':
        return HttpResponseForbidden("No tienes permiso para ver esta sección.")

    usuarios = User.objects.all().select_related("perfil").order_by("username")

    return render(request, "examenes/lista_usuarios.html", {"usuarios": usuarios})


def crear_usuario(request):
    # Solo el ADMIN puede crear usuarios
    perfil_admin = Perfil.objects.get(usuario=request.user)
    if perfil_admin.rol != "ADMIN":
        return HttpResponseForbidden("No tienes permiso para crear usuarios.")

    if request.method == "POST":
        form = UsuarioForm(request.POST)
        if form.is_valid():
            user = form.save()
            messages.success(request, f"✅ Usuario {user.username} creado correctamente.")
            return redirect("examenes:lista_usuarios")
    else:
        form = UsuarioForm()

    return render(request, "examenes/crear_usuario.html", {"form": form})


@login_required
def crear_usuario_ajax(request):
    perfil_admin = Perfil.objects.get(usuario=request.user)
    if perfil_admin.rol != "ADMIN":
        return JsonResponse({"error": "No tienes permiso para crear usuarios."}, status=403)

    if request.method == "POST":
        username = request.POST.get("username")
        email = request.POST.get("email")
        telefono = request.POST.get("telefono")
        rol = request.POST.get("rol")
        password = request.POST.get("password")

        # Validaciones básicas
        if not username or not password:
            return JsonResponse({"error": "El nombre de usuario y la contraseña son obligatorios"}, status=400)
        if User.objects.filter(username=username).exists():
            return JsonResponse({"error": "El usuario ya existe"}, status=400)

        # Crear usuario con contraseña
        user = User.objects.create_user(username=username, email=email, password=password)

        # Crear perfil asociado
        perfil = Perfil.objects.create(usuario=user, rol=rol, telefono=telefono)

        # Devolver JSON con todos los campos que js necesita
        return JsonResponse({
            "id": user.id,
            "username": user.username,
            "email": user.email,
            "telefono": perfil.telefono,
            "perfil__rol": perfil.rol,
        })

    return JsonResponse({"error": "Método no permitido"}, status=405)


@login_required
def editar_usuario(request, usuario_id):
    perfil_admin = Perfil.objects.get(usuario=request.user)
    if perfil_admin.rol != "ADMIN":
        return HttpResponseForbidden("No tienes permiso para editar usuarios.")

    usuario = get_object_or_404(User, id=usuario_id)

    if request.method == "POST":
        username = request.POST.get("username")
        email = request.POST.get("email")
        rol = request.POST.get("rol")
        telefono = request.POST.get("telefono")

        # Actualizar datos del User
        usuario.username = username
        usuario.email = email
        usuario.save()

        # Actualizar datos del Perfil
        usuario.perfil.rol = rol
        usuario.perfil.telefono = telefono
        usuario.perfil.save()

        messages.success(request, f"✅ El usuario {usuario.username} fue actualizado correctamente.")

        return redirect("examenes:lista_usuarios")

    return render(request, "examenes/editar_usuario.html", {"usuario": usuario})


@login_required
def eliminar_usuario(request, usuario_id):
    usuario = get_object_or_404(User, id=usuario_id)

    if request.method == "POST":
        nombre = usuario.username
        usuario.delete()
        messages.warning(request, f"🗑️ El usuario «{nombre}» fue eliminado correctamente.")
        return redirect('examenes:lista_usuarios')

    return render(request, "examenes/confirmar_eliminar_usuario.html", {"usuario": usuario})

@login_required
def crear_tipo_examen(request):
    perfil = Perfil.objects.get(usuario=request.user)
    if perfil.rol != 'ADMIN':
        return HttpResponseForbidden("No tienes permiso para crear tipos de examen.")

    if request.method == "POST":
        nombre = request.POST.get("nombre")
        unidad = request.POST.get("unidad")
        rango_bajo = request.POST.get("rango_bajo")
        rango_normal = request.POST.get("rango_normal")
        rango_alto = request.POST.get("rango_alto")

        TipoExamen.objects.create(
            nombre=nombre,
            unidad=unidad,
            rango_bajo=rango_bajo,
            rango_normal=rango_normal,
            rango_alto=rango_alto,
        )
        return redirect("examenes:ver_catalogo")

    return render(request, "examenes/crear_tipo_examen.html")
    catalogo = TipoExamen.objects.all()
    return render(request, "examenes/catalogo.html", {"catalogo": catalogo})

@login_required
def editar_tipo_examen(request, tipo_id):
    perfil = Perfil.objects.get(usuario=request.user)
    if perfil.rol != 'ADMIN':
        return HttpResponseForbidden("No tienes permiso para editar tipos de examen.")

    tipo = get_object_or_404(TipoExamen, id=tipo_id)

    if request.method == "POST":
        tipo.nombre = request.POST.get("nombre", tipo.nombre)
        tipo.unidad = request.POST.get("unidad", tipo.unidad)
        tipo.rango_bajo = request.POST.get("rango_bajo", tipo.rango_bajo)
        tipo.rango_normal = request.POST.get("rango_normal", tipo.rango_normal)
        tipo.rango_alto = request.POST.get("rango_alto", tipo.rango_alto)
        tipo.save()

        messages.success(request, f"✅ El examen clínico '{tipo.nombre}' fue editado correctamente.")

    return render(request, "examenes/editar_tipo_examen.html", {"tipo": tipo})

@login_required
def eliminar_tipo_examen(request, tipo_id):
    perfil = Perfil.objects.get(usuario=request.user)
    if perfil.rol != 'ADMIN':
        return HttpResponseForbidden("No tienes permiso para eliminar tipos de examen.")

    tipo = get_object_or_404(TipoExamen, id=tipo_id)

    # Solo eliminar si viene por POST
    if request.method == "POST":
        tipo.delete()
        messages.success(request, "🗑️ El examen clínico fue eliminado correctamente.")
        return redirect("examenes:ver_catalogo")

    # Si entra por GET, mostrar página de confirmación
    return redirect("examenes:ver_catalogo")



# --- ESTADÍSTICAS MÉDICO ---
@login_required
def estadisticas_medico(request):
    estadisticas = {
        'total_resultados': ResultadoExamen.objects.count(),
        'riesgo_alto': ResultadoExamen.objects.filter(nivel_riesgo='ALTO').count(),
        'riesgo_normal': ResultadoExamen.objects.filter(nivel_riesgo='NORMAL').count(),
        'riesgo_bajo': ResultadoExamen.objects.filter(nivel_riesgo='BAJO').count(),
    }
    return render(request, 'examenes/estadisticas.html', {'estadisticas': estadisticas})

# --- CATÁLOGO DE EXÁMENES ---
def ver_catalogo(request):
    catalogo = CatalogoExamen.objects.all()
    return render(request, "examenes/catalogo.html", {"catalogo": catalogo})


def crear_catalogo_examen(request):
    if request.method == 'POST':
        form = CatalogoExamenForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Examen registrado correctamente en el catálogo.")
            return redirect('examenes:catalogo')
        else:
            messages.error(request, "❌ Error al registrar el examen. Revisa los campos.")
    else:
        form = CatalogoExamenForm()
    return render(request, "examenes/crear_catalogo_examen.html", {"form": form})


def eliminar_catalogo_examen(request, id):
    examen = get_object_or_404(CatalogoExamen, id=id)
    examen.delete()
    messages.warning(request, f"🗑️ El examen «{examen.nombre}» fue eliminado del catálogo.")
    return redirect('examenes:catalogo')

def editar_catalogo_examen(request, id):
    examen = get_object_or_404(CatalogoExamen, id=id)

    if request.method == 'POST':
        form = CatalogoExamenForm(request.POST, instance=examen)
        if form.is_valid():
            form.save()
            messages.success(request, f"✅ El examen clínico '{examen.nombre}' fue editado correctamente.")
            # Redirigir al catálogo
            return redirect('examenes:catalogo')
    else:
        form = CatalogoExamenForm(instance=examen)

    return render(request, "examenes/editar_catalogo_examen.html", {"form": form})


# --- GRÁFICO DE GLUCOSA ---

@login_required
def grafico_glucosa(request):
    # Filtra los resultados del paciente actual
    examenes = ResultadoExamen.objects.filter(paciente=request.user).select_related("examen_tipo").order_by("fecha_examen")

    data_por_examen = {}

    for ex in examenes:
        nombre = ex.examen_tipo.nombre

        if nombre not in data_por_examen:
            data_por_examen[nombre] = {"labels": [], "values": []}

        # Fecha como etiqueta
        fecha = ex.fecha_examen.strftime("%d/%m/%Y") if ex.fecha_examen else ""

        # Agrega el punto con su rango clínico
        data_por_examen[nombre]["labels"].append(fecha)
        data_por_examen[nombre]["values"].append({
            "valor": float(ex.valor),
            "rango_min": float(ex.examen_tipo.rango_min),
            "rango_max": float(ex.examen_tipo.rango_max)
        })

    return render(request, "examenes/graficos.html", {
        "data_json": json.dumps(data_por_examen)
    })


# --- DASHBOARD GENERAL ---
@login_required
def dashboard(request):
    examenes = ResultadoExamen.objects.filter(paciente=request.user).order_by('-fecha_examen')
    tipos = ["Glucosa", "Colesterol", "Triglicéridos", "Diabetes"]
    graficos = {}
    for tipo in tipos:
        datos = examenes.filter(examen_tipo__nombre=tipo).order_by("fecha_examen")
        fechas = [ex.fecha_examen.strftime("%d/%m/%Y") for ex in datos]
        valores = [float(ex.valor) for ex in datos]
        graficos[tipo] = {"fechas": fechas, "valores": valores}
    return render(request, "examenes/dashboard.html", {
        "examenes": examenes,
        "graficos_json": json.dumps(graficos),
    })

# --- EXPORTAR EXCEL ---
@login_required
def exportar_reporte_excel (request, paciente_id=None):
    perfil, _ = Perfil.objects.get_or_create(usuario=request.user, defaults={'rol': 'PACIENTE'})
    if perfil.rol == 'PACIENTE':
        examenes = ResultadoExamen.objects.filter(paciente=request.user).order_by("fecha_examen")
    else:
        examenes = ResultadoExamen.objects.filter(paciente_id=paciente_id).order_by("fecha_examen")

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Exámenes"

    headers = ["Fecha", "Tipo de Examen", "Valor", "Nivel de Riesgo"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for examen in examenes:
        ws.append([
            examen.fecha_examen.strftime("%d/%m/%Y"),
            examen.examen_tipo.nombre,
            examen.valor,
            examen.nivel_riesgo,
        ])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="reporte_paciente_{paciente_id or request.user.id}.xlsx"'
    wb.save(response)
    return response

# --- API REST ---
class ExamenViewSet(viewsets.ModelViewSet):
    queryset = ResultadoExamen.objects.all()
    serializer_class = ExamenSerializer
    filter_backends = [drf_filters.SearchFilter]
    filterset_fields = ["nivel_riesgo", "examen_tipo__nombre", "fecha_examen"]
    search_fields = ["paciente__username", "examen_tipo__nombre"]

    @action(detail=False, methods=["post"], url_path="carga-masiva")
    def carga_masiva(self, request):
        archivo = request.FILES.get("file")
        if not archivo:
            return Response({"error": "No se envió ningún archivo"}, status=status.HTTP_400_BAD_REQUEST)

        wb = openpyxl.load_workbook(archivo)
        ws = wb.active
        creados = 0
        errores = []
        for fila in ws.iter_rows(min_row=2, values_only=True):
            try:
                fecha, paciente_id, tipo_id, valor, riesgo = fila
                ResultadoExamen.objects.create(
                    paciente_id=paciente_id,
                    examen_tipo_id=tipo_id,
                    fecha_examen=fecha,
                    valor=valor,
                    nivel_riesgo=riesgo,
                )
                creados += 1
            except Exception as e:
                errores.append(str(e))

        return Response({
            "mensaje": f"Se cargaron {creados} registros",
            "errores": errores
        }, status=status.HTTP_201_CREATED)